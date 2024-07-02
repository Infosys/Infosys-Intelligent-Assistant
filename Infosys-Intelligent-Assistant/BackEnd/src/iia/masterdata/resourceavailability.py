__created__ = "Mar 19, 2019"
__copyright__ = """Copyright 2022 Infosys Ltd.
Use of this source code is governed by MIT license that can be found in the LICENSE file or at
https://opensource.org/licenses/MIT."""

from iia.persistence.mongodbpersistence import MongoDBPersistence
from iia.restservice import RestService
from datetime import datetime, date
from flask import request
from bson import json_util
import re
import xlrd
from iia.utils.log_helper import get_logger
import yaml
import openpyxl
import pandas as pd
import pyexcel as p

logging = get_logger(__name__)
app = RestService.getApp();


@app.route('/api/uploadRoasterDetails/<int:customer_id>/<team_name>', methods=["POST"])
def db_insert_roaster_details(customer_id, team_name):
    return ResourceAvailabilityMasterData.db_insert_roaster_details(customer_id, team_name)


@app.route('/api/getShiftDetails/<int:customer_id>/<team_name>/<email_id>', methods=['GET'])
def get_roaster_details_of_resource(customer_id, team_name, email_id):
    return ResourceAvailabilityMasterData.get_roaster_details_of_resource(customer_id, team_name, email_id)


@app.route('/api/updateRoaster/<int:customer_id>/<team_name>/<email_id>/<start_date>/<end_date>', methods=['PUT'])
def update_roaster(customer_id, team_name, email_id, start_date, end_date):
    return ResourceAvailabilityMasterData.update_roaster(customer_id, team_name, email_id, start_date, end_date)


@app.route('/api/getRoasterDetails/<int:customer_id>/<team_name>', methods=['GET'])
def get_roaster_details(customer_id, team_name):
    return ResourceAvailabilityMasterData.get_roaster_details(customer_id, team_name)


@app.route('/api/getRoasterTeamNames', methods=['GET'])
def get_roaster_team_names():
    return ResourceAvailabilityMasterData.get_roaster_team_names()


class ResourceAvailabilityMasterData(object):
    '''
    classdocs
    '''

    def __init__(self, params):
        '''
        Constructor
        '''

    def getRoasterList(assign_grp):

        logging.info('%s: In "getRoasterList()" method' % RestService.timestamp())
        app_name = '.*'
        roaster_lst = []

        # --code to get today's date edited today--
        logging.info('Trying to get todays date and time')
        date1 = date.today()
        time1 = datetime.now().time()
        today = str(date1) + ' T' + str(format(time1.hour, '02d')) + ':' + str(format(time1.minute, '02d')) + ':' + str(
            format(time1.second, '02d'))

        logging.info('trying to get documents from "roaster_tbl" for today: %s' % today)
        roaster_lst = list(MongoDBPersistence.roaster_tbl.aggregate([
            {
                '$match': {
                    '$and': [
                        {'start_date': {'$lte': today}},
                        {'end_date': {'$gte': today}}
                    ], 'availability': True, 'assignment_group': assign_grp
                }
            }, {
                '$lookup': {
                    'from': "TblResource",
                    'localField': "email_id",
                    'foreignField': "email_id",
                    'as': "resource_name"
                }
            }, {
                '$project': {
                    '_id': 0, 'resource_name.resource_name': 1, 'shift': 1, 'support_type': 1, 'email_id': 1
                }
            }

        ]))

        if (len(roaster_lst) == 0):
            logging.error('Could not get documents from "roaster_tbl"')

        return roaster_lst

    @staticmethod
    def db_insert_roaster_details(customer_id, team_name):
        print('\nInside db_insert_roaster_details for excel upload')
        try:

            global_start_date_current_month = ""
            global_end_date_current_month = ""
            start_time = datetime.now()
            logging.info(
                '%s: In "db_insert_roaster_details()" method, Trying to get file from request object' % RestService.timestamp())
            file = request.files['roasterDetails']
            # print('type of file: ',type(file))
            dataset_ = MongoDBPersistence.teams_tbl.find_one({"CustomerID": customer_id, "TeamName": team_name},
                                                             {"DatasetID": 1, "_id": 0})
            if (dataset_):
                # Dataset exist for the team
                logging.info('%s: Getting old dataset details.' % RestService.timestamp())
                dataset_id = dataset_["DatasetID"]
            else:
                # Newly adding the dataset
                logging.info('%s: Adding new dataset.' % RestService.timestamp())
                # getting max dataset id for the customer, so that new dataset id = old + 1
                dataset_dict = MongoDBPersistence.datasets_tbl.find_one(
                    {"CustomerID": customer_id, "DatasetID": {"$exists": True}},
                    {'_id': 0, "DatasetID": 1, "DatasetName": 1}, sort=[("DatasetID", -1)])
                if (dataset_dict):
                    last_dataset_id = dataset_dict['DatasetID']
                    dataset_name = dataset_dict['DatasetName']
                else:
                    last_dataset_id = 0
                    dataset_name = team_name
                    logging.info('%s: Adding dataset for very first team.' % RestService.timestamp())

                # New dataset id for the customer
                dataset_id = last_dataset_id + 1

                new_dataset_dict = {}
                new_dataset_dict["DatasetID"] = dataset_id
                new_dataset_dict["DatasetName"] = dataset_name
                new_dataset_dict["CustomerID"] = customer_id

                MongoDBPersistence.datasets_tbl.insert_one(new_dataset_dict)

            if not file:
                logging.info('%s:No file' % RestService.timestamp())
                return json_util.dumps({'status': 'failure'})
            elif (not '.xls' in str(file)):
                logging.info('%s: Uploded file is not excel, insert excel file to continue' % RestService.timestamp())
                return json_util.dumps({'status': 'failure'})
            fileName = "./data/temp_" + team_name + ".xls"
            file.save(fileName)

            fileName_xlsx = "./data/temp_" + team_name + ".xlsx"
            p.save_book_as(file_name=fileName, dest_file_name=fileName_xlsx)

            # --fetching data from excel sheet--
            wb1 = xlrd.open_workbook(fileName)

            wb = openpyxl.open(fileName_xlsx, read_only=True)

            #    --Validationsss--
            no_of_sheets = len(wb1.sheet_names())
            resource_lst = []
            app_lst = []
            email_lst = []
            assignment_lst = []
            ignored_lst = []

            email_id_in_roaster = []
            holiday = {}

            shift_dates_list = []
        except Exception as e:
            print("Exception occurred while reading or uploading the excel : ", e)
            logging.info("Exception occurred while reading or uploading the excel : %s" % e)
            return "failure"
        try:
            logging.info('Trying to get documents from "resource_details_tbl", "applicationDetails_tbl"')
            resource_lst = list(
                MongoDBPersistence.resource_details_tbl.find({'CustomerID': customer_id, 'DatasetID': dataset_id},
                                                             {'_id': 0, 'email_id': 1}))
            app_lst = list(
                MongoDBPersistence.applicationDetails_tbl.find({'CustomerID': customer_id, 'DatasetID': dataset_id},
                                                               {'_id': 0, 'assignment_group': 1}))

            with open('./config/shift_roaster_config.yaml') as roaster_file:
                roaster_config = yaml.safe_load(roaster_file)

            df_shift_roaster = pd.DataFrame(roaster_config).transpose()
            df_shift_roaster = df_shift_roaster.reset_index()
            df_shift_roaster = df_shift_roaster.rename(columns={'index': 'shift'})

            if (len(resource_lst) != 0 or len(app_lst) != 0):

                for resource_doc in resource_lst:
                    email_lst.append(resource_doc['email_id'])
                for app_doc in app_lst:
                    assignment_lst.append(app_doc['assignment_group'])

                master_df = None
                for ws in wb.worksheets:
                    value = ws["M2"].value
                    if value == "Month":
                        sheet_name = ws.title
                        assignment_group = ws["C2"].value
                        year_month = ws["P2"].value
                        application_segment_1 = ws["C3"].value
                        application_segment_2 = ws["O2"].value
                        application_name = ws["AA3"].value
                        df = pd.read_excel(fileName, skiprows=5, sheet_name=sheet_name)
                        df = df.melt(id_vars='resource_id')
                        df = df.rename(columns={"resource_id": 'email_id', "variable": 'day', "value": 'shift'})
                        df['day'] = df['day'].astype(str).apply(lambda x: x.zfill(2))
                        df['year_month'] = year_month
                        df['date'] = df['year_month'] + '-' + df['day'].astype(str)
                        df['assignment_group'] = assignment_group
                        df['application_name'] = application_name
                        df['CustomerID'] = 1
                        df['DatasetID'] = 1
                        df['availability'] = True
                        df = df[['CustomerID', 'DatasetID', 'assignment_group', 'email_id', 'application_name', 'shift',
                                 'date', 'availability']]
                        df['date'] = df['date'].astype(str)
                        try:
                            if master_df is None:
                                master_df = df
                            else:
                                master_df = pd.concat([master_df, df], axis=0, ignore_index=True)
                        except Exception as e:
                            print(e)

                filtered_df = master_df
                filtered_df = filtered_df.drop_duplicates(
                    subset=['assignment_group', 'email_id', 'application_name', 'date'], keep='last').reset_index(
                    drop=True)
                filtered_df = pd.merge(filtered_df, df_shift_roaster, on='shift')
                filtered_df['start_date'] = filtered_df['date'] + ' T' + filtered_df['start_time']
                filtered_df['end_date'] = filtered_df['date'] + ' T' + filtered_df['end_time']
                year_month = year_month[-2:]
                month_dict = {'01': 'January', '02': 'February', '03': 'March', '04': 'April', '05': 'May',
                              '06': 'June',
                              '07': 'July', '08': 'August', '09': 'September', '10': 'October', '11': 'November',
                              '12': 'December'}

                filtered_df['Month'] = month_dict[year_month]

                filtered_df = filtered_df[
                    ['CustomerID', 'DatasetID', 'assignment_group', 'email_id', 'application_name', 'shift',
                     'start_date',
                     'end_date', 'availability', 'support_type', 'Month']]
                df_records = filtered_df.to_dict('records')
                
                # If roaster is already present in db do not add duplicate data again
                for item in df_records:
                    k = list(MongoDBPersistence.roaster_tbl.find({
                        "email_id": item['email_id'], "DatasetID": item['DatasetID'],
                        "start_date": item['start_date'], "end_date": item['end_date'],
                        "Month": item['Month'], "application_name": item['application_name'],
                        "shift": item['shift'], "assignment_group": item['assignment_group']}))
                    if (len(k) == 0):
                        MongoDBPersistence.roaster_tbl.insert_one({
                            "CustomerID": item['CustomerID'], "DatasetID": item['DatasetID'],
                            "assignment_group": item['assignment_group'],
                            "email_id": item['email_id'], "application_name": item['application_name'],
                            "shift": item['shift'], "start_date": item['start_date'],
                            "end_date": item['end_date'], "availability": item['availability'],
                            "support_type": item['support_type'], "Month": item['Month']})
                    else:
                        pass

                shift_values = ['H', 'L', 'C']

                hlc = master_df.loc[master_df['shift'].isin(shift_values)]

                for index, row in master_df.iterrows():
                    email_id = row["email_id"]
                    if (email_id not in email_id_in_roaster):
                        email_id_in_roaster.append(email_id)

                for index, row in hlc.iterrows():
                    email_id = row["email_id"]
                    date = row["date"]
                    if (email_id not in holiday):
                        holiday[email_id] = list()
                        holiday[email_id].append(date[0:4] + "/" + date[5:7] + "/" + date[8:10])
                    else:
                        if (date not in holiday[email_id]):
                            holiday[email_id].append(date[0:4] + "/" + date[5:7] + "/" + date[8:10])
            else:
                logging.error(
                    'Could not get documents from either "resource_details_tbl" or from "applicationDetails_tbl" for cutomer id: %s, dataset id: %s' % (
                    str(customer_id), str(dataset_id)))
        except Exception as e:
            logging.error('Error: %s' % str(e))
            return json_util.dumps({'status': 'failure'})

        try:
            print(" email_id_in_roaster--------------------", email_id_in_roaster)
            dup = list(MongoDBPersistence.application_analyst_mapping_tbl.find({}, {"_id": 0, "email_id": 1}))
            application_analyst_resource_email = []
            for i in dup:
                application_analyst_resource_email.append(i['email_id'])
            print("application_analyst_resource_email", application_analyst_resource_email)
            print(
                '******************************************* END OF EMAIL ID *******************************************')
            for i in email_id_in_roaster:

                logging.info(f"Updating Roaser for {i}")
                
                roaster_assign_grp = MongoDBPersistence.roaster_tbl.distinct("assignment_group", {"email_id": i})
                logging.info(f"roaster_assign_grp: {roaster_assign_grp}")
                resource_id_name = list(MongoDBPersistence.resource_details_tbl.find({"email_id": i},
                                                                                     {"_id": 0, "resource_id": 1,
                                                                                      "resource_name": 1,
                                                                                      "email_id": 1}))
                logging.info(f"resource_id_name: {resource_id_name}")
                resource_id = resource_id_name[0]['resource_id']
                resource_name = resource_id_name[0]['resource_name']
                email_id = resource_id_name[0]['email_id']
                logging.info(f"resource_name: {resource_name}")
                logging.info(f"email_id: {email_id}")
               
                MongoDBPersistence.application_analyst_mapping_tbl.update_one({"resource_id": resource_id}, {
                    "$set": {"resource_id": resource_id, "email_id": i,
                             "resource_group": roaster_assign_grp, "resource_name": resource_name,
                             "DatasetID": dataset_id}}, upsert=True)

                sheet = wb1.sheet_by_index(1)
                year_month = sheet.cell_value(1, 15)
                month = int(year_month[5:7])

                for j in roaster_assign_grp:
                    startdate_value = 0
                    if (month in [1, 3, 5, 7, 8, 10, 12]):
                        shift_list = [""] * 31
                    elif (month in [4, 6, 9, 11]):
                        shift_list = [""] * 30
                    else:
                        shift_list = [""] * 28
                    temp_date = []
                    shift_name = []

                    year_month = year_month[-2:]
                    current_month = month_dict[year_month]
                    
                    op = list(MongoDBPersistence.roaster_tbl.find(
                        {"email_id": email_id, "assignment_group": j, "Month": current_month},
                        {"_id": 0, "start_date": 1, "end_date": 1, "shift": 1}))
                    
                    st_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][5:7] + "/" + op[0]["start_date"][
                                                                                                8:10]
                    ed_date = op[-1]["start_date"][0:4] + "/" + op[-1]["start_date"][5:7] + "/" + op[-1]["start_date"][
                                                                                                  8:10]
                    global_start_date_current_month = st_date
                    global_end_date_current_month = ed_date

                    for o in op:
                        shift_name.append(o['shift'])
                    
                    year_month = year_month[-2:]
                   
                    search_month = month_dict[year_month]
                    shifts = list(MongoDBPersistence.roaster_tbl.find(
                        {"email_id": email_id, "assignment_group": j, "Month": search_month},
                        {"_id": 0, "start_date": 1}))
                    print(
                        '******************************************* SHIFT DETAILS START *******************************************')
                    
                    print(
                        '******************************************* SHIFT DETAILS END *******************************************')

                    for date in shifts:
                        act_date = int(date["start_date"][8:10])
                        if (act_date < 10):
                            temp_date.append("0" + str(act_date))
                        else:
                            temp_date.append(act_date)
                    
                    for date_s in temp_date:
                        shift_list[int(date_s) - 1] = shift_name[startdate_value]
                        startdate_value = startdate_value + 1

                    count = 0

                    temp_end_date = 0

                    for sh in range(0, len(shift_list)):
                        # 1.last date of month
                        if (sh == len(shift_list) - 1 and shift_list[sh] != ""):

                            temp_end_date = sh + 1
                            temp_end_date_for_last = "01"
                            temp_st_date = sh + 1
                            temp_shift = shift_list[sh]
                            # 1a.if shifts are not same
                            if (temp_shift != shift_list[sh - 1]):
                                year_value = str(int(op[0]["start_date"][0:4]) + 1)
                                month_value = int(op[0]["start_date"][5:7])
                                next_month_value = month_value + 1
                                if (next_month_value < 10):
                                    next_month_value = "0" + str(next_month_value)
                                if (int(month_value) == 12):
                                    formated_end_date = year_value + "/" + str(temp_end_date_for_last) + "/" + str(
                                        temp_end_date_for_last)
                                else:
                                    formated_end_date = op[0]["start_date"][0:4] + "/" + str(
                                        next_month_value) + "/" + str(temp_end_date_for_last)
                                formated_st_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][
                                                                                    5:7] + "/" + str(temp_st_date)

                                repeat = list(MongoDBPersistence.roster_mapping_tbl.find({
                                    "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                    "start_date": formated_st_date, "end_date": formated_end_date,
                                    "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                    "resource_name": resource_name, "temp_resource_id": resource_id,
                                    "resource_shift": temp_shift}))

                                if len(repeat) == 0:
                                    MongoDBPersistence.roster_mapping_tbl.insert_one({
                                        "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                        "start_date": formated_st_date, "end_date": formated_end_date,
                                        "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                        "resource_name": resource_name, "temp_resource_id": resource_id,
                                        "resource_shift": temp_shift})

                        # 2.if same shift
                        if (sh < temp_end_date and sh != 0):
                            continue
                        ind = sh + 1

                        # 3.if shift is not ""
                        if (shift_list[sh] != ""):
                            temp_shift = shift_list[sh]
                            temp_st_date = ind
                            for sha in range(ind, len(shift_list) + 1):
                                # 3a-last shift
                                if (sha == len(shift_list)):

                                    temp_end_date = sha
                                    if (temp_end_date < 10):
                                        temp_end_date = "0" + str(temp_end_date)
                                    formated_end_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][
                                                                                         5:7] + "/" + str(temp_end_date)
                                    if (temp_st_date < 10):
                                        temp_st_date = "0" + str(temp_st_date)

                                    formated_st_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][
                                                                                        5:7] + "/" + str(temp_st_date)
                                    temp_st_date = int(temp_st_date)
                                    temp_end_date = int(temp_end_date)
                                    
                                    repeat = list(MongoDBPersistence.roster_mapping_tbl.find({
                                        "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                        "start_date": formated_st_date, "end_date": formated_end_date,
                                        "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                        "resource_name": resource_name, "temp_resource_id": resource_id,
                                        "resource_shift": temp_shift}))

                                    if len(repeat) == 0:
                                        MongoDBPersistence.roster_mapping_tbl.insert_one({
                                            "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                            "start_date": formated_st_date, "end_date": formated_end_date,
                                            "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                            "resource_name": resource_name, "temp_resource_id": resource_id,
                                            "resource_shift": temp_shift})


                                # 3b-nesxt shift value = ""
                                elif (shift_list[sha] == ""):

                                    if (sha + 1 < len(shift_list)):
                                        if (shift_list[sha + 1] != ""):
                                            temp_end_date = sha + 1

                                            if (temp_end_date < 10):
                                                temp_end_date = "0" + str(temp_end_date)

                                            formated_end_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][
                                                                                                 5:7] + "/" + str(
                                                temp_end_date)
                                            if (temp_st_date < 10):
                                                temp_st_date = "0" + str(temp_st_date)
                                            formated_st_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][
                                                                                                5:7] + "/" + str(
                                                temp_st_date)
                                            temp_st_date = int(temp_st_date)
                                            temp_end_date = int(temp_end_date)
                                            
                                            repeat = list(MongoDBPersistence.roster_mapping_tbl.find({
                                                "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                                "start_date": formated_st_date, "end_date": formated_end_date,
                                                "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                                "resource_name": resource_name, "temp_resource_id": resource_id,
                                                "resource_shift": temp_shift}))
                                            if len(repeat) == 0:
                                                MongoDBPersistence.roster_mapping_tbl.insert_one({
                                                    "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                                    "start_date": formated_st_date, "end_date": formated_end_date,
                                                    "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                                    "resource_name": resource_name, "temp_resource_id": resource_id,
                                                    "resource_shift": temp_shift})
                                            break

                                        else:
                                            if (sha == len(shift_list)):
                                                temp_end_date = sha
                                                formated_end_date = op[0]["start_date"][0:4] + "/" + op[0][
                                                                                                         "start_date"][
                                                                                                     5:7] + "/" + str(
                                                    temp_end_date)
                                                formated_st_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][
                                                                                                    5:7] + "/" + str(
                                                    temp_st_date)
                                                
                                                temp_st_date = int(temp_st_date)
                                                temp_end_date = int(temp_end_date)

                                                repeat = list(MongoDBPersistence.roster_mapping_tbl.find({
                                                    "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                                    "start_date": formated_st_date, "end_date": formated_end_date,
                                                    "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                                    "resource_name": resource_name, "temp_resource_id": resource_id,
                                                    "resource_shift": temp_shift}))

                                                if len(repeat) == 0:
                                                    MongoDBPersistence.roster_mapping_tbl.insert_one({
                                                        "resource_id": resource_id, "email_id": i,
                                                        "DatasetID": dataset_id,
                                                        "start_date": formated_st_date, "end_date": formated_end_date,
                                                        "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                                        "resource_name": resource_name, "temp_resource_id": resource_id,
                                                        "resource_shift": temp_shift})

                                            continue

                                # 4.if current_shift == previousshift
                                elif (shift_list[sha] == temp_shift):
                                    continue

                                # 5.if current_shift != previousshift
                                elif (shift_list[sha] != temp_shift):
                                    temp_end_date = sha

                                    if (temp_end_date < 10):
                                        temp_end_date = "0" + str(temp_end_date)

                                    formated_end_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][
                                                                                         5:7] + "/" + str(temp_end_date)
        
                                    if (temp_st_date < 10):
                                        temp_st_date = "0" + str(temp_st_date)

                                    formated_st_date = op[0]["start_date"][0:4] + "/" + op[0]["start_date"][
                                                                                        5:7] + "/" + str(temp_st_date)
                                    temp_st_date = int(temp_st_date)
                                    temp_end_date = int(temp_end_date)

                                    repeat = list(MongoDBPersistence.roster_mapping_tbl.find({
                                        "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                        "start_date": formated_st_date, "end_date": formated_end_date,
                                        "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                        "resource_name": resource_name, "temp_resource_id": resource_id,
                                        "resource_shift": temp_shift}))
                                    if len(repeat) == 0:
                                        MongoDBPersistence.roster_mapping_tbl.insert_one({
                                            "resource_id": resource_id, "email_id": i, "DatasetID": dataset_id,
                                            "start_date": formated_st_date, "end_date": formated_end_date,
                                            "list_shift": shift_list, "resource_group": roaster_assign_grp,
                                            "resource_name": resource_name, "temp_resource_id": resource_id,
                                            "resource_shift": temp_shift})

                                    break
                        else:
                            continue
                    
                    resource_roastermap = list(
                        MongoDBPersistence.roster_mapping_tbl.find({"email_id": i, "resource_group": j,
                                                                    '$and': [{"start_date": {
                                                                        "$gte": global_start_date_current_month}}]},
                                                                   {"email_id": 1, "start_date": 1, "end_date": 1}))
                    
                    if (holiday):
                        if i not in holiday.keys():
                            holiday[i] = []
                        holiday_of_resource = holiday[i]
                        for hol in range(len(resource_roastermap)):
                            name = resource_roastermap[hol]["email_id"]
                            new_hol_list = []
                            hol_ad = hol + 1
                            hol_start_date = int(resource_roastermap[hol]["start_date"][8:])
                            hol_end_date = int(resource_roastermap[hol]["end_date"][8:])

                            for h_res in holiday_of_resource:
                                if (int(h_res[8:]) >= hol_start_date and int(
                                        h_res[8:]) <= hol_end_date and h_res not in new_hol_list):
                                    new_hol_list.append(h_res)
                           
                            year_month = year_month[-2:]

                            current_month = month_dict[year_month]

                            MongoDBPersistence.roster_mapping_tbl.update_one({
                                "email_id": i,
                                "end_date": resource_roastermap[hol]["end_date"]},
                                {"$set": {"weekoff": new_hol_list}})

                    print(
                        "********************************************** Holiday's are updated **********************************************")

            print('response db_insert_roaster_details Success')
            logging.info('Roaster successfully uploaded!, returning ignored list if any')
            return "Success"
        except Exception as e:
            print("exception occurred while saving the data", e)
            logging.info("Exception in insertDB shift details function%s" % e)
            print('response insertDBShiftDetails failure')
            logging.info('Roaster successfully uploaded!, returning ignored list if any')
            return "failure"
           
    @staticmethod
    def get_roaster_details_of_resource(customer_id, team_name, email_id):
        try:

            logging.info('%s: Trying to get document from "teams_tbl"' % RestService.timestamp())
            team_doc = MongoDBPersistence.teams_tbl.find_one({'TeamName': team_name, 'DatasetID': {'$exists': True}},
                                                             {'_id': 0, 'DatasetID': 1})

            if (team_doc != None):
                dataset_id = team_doc['DatasetID']
                logging.info('%s: Trying to get document from "roaster_tbl".' % RestService.timestamp())
                roaster_lst = list(MongoDBPersistence.roaster_tbl.find({
                    'email_id': email_id,
                    'availability': True, 'CustomerID': customer_id,
                    'DatasetID': dataset_id
                }, {
                    '_id': 0, 'assignment_group': 1, 'shift': 1,
                    'support_type': 1, 'start_date': 1, 'availability': 1
                }))

                if (len(roaster_lst) != 0):
                    for roaster_doc in roaster_lst:
                        roaster_doc['start_date'] = roaster_doc['start_date'].split(' ')[0]
                else:
                    logging.error('Could not get documents from "roaster_tbl": for email: %s' % email_id)
            else:
                logging.error('Could not get document from "teams_tbl" for team: %s' % team_name)

        except Exception as e:
            logging.info('Error: %s', str(e))
            return ('failure')

        return (json_util.dumps(roaster_lst))

    @staticmethod
    def update_roaster(customer_id, team_name, email_id, start_date, end_date):

        logging.info('%s: In "update_roaster()" method' % RestService.timestamp())
        a = {'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
             'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'}
        print("start_date..........", start_date, int(start_date))
        date1 = datetime.fromtimestamp(int(start_date) / 1000)
        print("date1...............", date1)
        day = date1.day
        month = date1.month
        if (len(str(day)) == 1):
            day = '0' + str(day)
        if (len(str(month)) == 1):
            month = '0' + str(month)
        start_date = str(date1.year) + '-' + str(month) + '-' + str(day) + ' T00:00:00'
        date2 = datetime.fromtimestamp(int(end_date) / 1000)
        
        day = date2.day
        month = date2.month
        if (len(str(day)) == 1):
            day = '0' + str(day)
        if (len(str(month)) == 1):
            month = '0' + str(month)
        end_date = str(date2.year) + '-' + str(month) + '-' + str(day) + ' T24:00:00'

        try:
            logging.info('Trying to get document from "teams_tbl" for team: %s' % team_name)
            team_doc = MongoDBPersistence.teams_tbl.find_one({'TeamName': team_name, 'DatasetID': {'$exists': True}},
                                                             {'_id': 0, 'DatasetID': 1})

            if (team_doc != None):
                resourcedetails = list(MongoDBPersistence.resource_details_tbl.find({"email_id": email_id},
                                                                                    {"_id": 0, "resource_name": 1,
                                                                                     "resource_id": 1}))
                dataset_id = team_doc['DatasetID']
                MongoDBPersistence.roaster_tbl.update_many({
                    'email_id': email_id, '$and': [{
                        'start_date': {'$gte': start_date}}, {
                        'end_date': {'$lte': end_date}
                    }], 'CustomerID': customer_id, 'DatasetID': dataset_id
                }, {'$set': {'availability': False, "shift": "L"}})

                print("insert successfull into leave details")

                day = date1.day
                month = date1.month
                if (len(str(day)) == 1):
                    day = '0' + str(day)
                if (len(str(month)) == 1):
                    month = '0' + str(month)
                mappingstart_date = str(date1.year) + '/' + str(month) + '/' + str(day)

                day = date2.day
                month = date2.month
                if (len(str(day)) == 1):
                    day = '0' + str(day)
                if (len(str(month)) == 1):
                    month = '0' + str(month)
                mappingend_date = str(date2.year) + '/' + str(month) + '/' + str(day)
                MongoDBPersistence.roster_mapping_tbl.update_one({
                    'email_id': email_id,
                    'start_date': mappingstart_date,
                    'end_date': mappingend_date
                }, {'$set': {"resource_shift": "L", "weekoff": [], "list_shift": [], "email_id": email_id,
                             'start_date': mappingstart_date,
                             'end_date': mappingend_date, "resource_name": resourcedetails[0]["resource_name"],
                             "resource_id": resourcedetails[0]["resource_id"]}}, upsert=True)
            else:
                logging.error('Could not get document from "teams_tbl"')
        except Exception as e:
            logging.error('Error: %s' % str(e))
            return ('failure')

        return ('success')

    @staticmethod
    def get_roaster_details(customer_id, team_name):

        assign_grp = request.args.get('assignment_group')
        today = request.args.get('date')

        response_doc = {}
        roaster_lst = []
        response_lst = []

        if (assign_grp == "" or assign_grp is None):
            assign_grp = ".*"

        try:
            if (today == None):
                todays_date = date.today()
                time = datetime.now().time()
                logging.info('%s: formatting todays date and current time' % RestService.timestamp())
                today = str(todays_date) + ' T' + str(format(time.hour, '02d')) + ':' + str(time.minute) + ':' + str(
                    time.second)
            logging.info(
                '%s: formated date is: %s, Trying to get document from "teams_tbl"' % (RestService.timestamp(), today))
            team_doc = MongoDBPersistence.teams_tbl.find_one({'TeamName': team_name, 'DatasetID': {'$exists': True}},
                                                             {'_id': 0, 'DatasetID': 1})

            if (team_doc != None):
                dataset_id = team_doc['DatasetID']

                roaster_lst = list(MongoDBPersistence.roaster_tbl.aggregate([
                    {
                        '$match': {
                            '$and': [
                                {'start_date': {'$lte': today}},
                                {'end_date': {'$gte': today}},
                                {'assignment_group': re.compile(assign_grp)}
                            ], 'CustomerID': customer_id,
                            'DatasetID': dataset_id, 'availability': True
                        }
                    }, {
                        '$lookup': {
                            'from': "TblResource",
                            'localField': "email_id",
                            'foreignField': "email_id",
                            'as': "resource_name"
                        }
                    }, {
                        '$project': {
                            '_id': 0, 'resource_name.resource_name': 1,
                            'shift': 1, 'support_type': 1, 'email_id': 1,
                            'start_date': 1, 'end_date': 1
                        }
                    }

                ]))
               
                if (len(roaster_lst) != 0):
                    for roaster_doc in roaster_lst:
                        response_doc['resource_name'] = roaster_doc['resource_name'][0]['resource_name']
                        response_doc['shift'] = roaster_doc['shift']
                        response_doc['support_type'] = roaster_doc['support_type']
                        response_doc['email_id'] = roaster_doc['email_id']
                        response_doc['start_date'] = roaster_doc['start_date'].split('T')[1]
                        response_doc['end_date'] = roaster_doc['end_date'].split('T')[1]
                        response_lst.append(response_doc)
                        response_doc = {}
                elif (MongoDBPersistence.roaster_tbl.find_one({})):
                    roaster_lst = MongoDBPersistence.roaster_tbl.find_one({'start_date': {'$gte': today}})
                    if (roaster_lst):
                        logging.error('%s: Roaster list is empty for the current shift' % RestService.timestamp())
                        return json_util.dumps('failure-01')
                    else:
                        today = todays_date.year + '-' + str(format(todays_date.month)) + '-01 T00:00:00'
                        roaster_lst = MongoDBPersistence.roaster_tbl.find_one({'start_date': {'$gte': today}})
                        logging.error(
                            '%s:  No further data for the current month or No data for the current month in "roaster_tbl"' % RestService.timestamp())
                        if (not roaster_lst):
                            logging.error('%s: Could not get shift details for future!' % RestService.timestamp())
                            return json_util.dumps('failure-02')
                        else:
                            logging.error(
                                '%s: Could not get shift details for further days of the month: %s from "roaster_tbl"!' % (
                                RestService.timestamp(), str(todays_date.month)))
                            return json_util.dumps('failure-03')
            else:
                logging.error('Could not get document from "teams_tbl" for team: %s' % team_name)

        except Exception as e:
            logging.error('Error: %s' % str(e))

        return (json_util.dumps(response_lst))

    @staticmethod
    def get_roaster_team_names():

        resource_doc = MongoDBPersistence.roaster_tbl.find_one({}, {'_id': 0, 'CustomerID': 1, 'DatasetID': 1})

        if (resource_doc):
            team = MongoDBPersistence.teams_tbl.find_one(
                {"CustomerID": resource_doc['CustomerID'], "DatasetID": resource_doc['DatasetID']},
                {'_id': 0, 'TeamName': 1})['TeamName']
            return (team);
        logging.info("no team")
        return ('no team')
